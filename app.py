import streamlit as st
import pandas as pd
import re
from openpyxl import load_workbook
import openpyxl.cell.cell
from io import BytesIO
import tempfile
import os

st.set_page_config(page_title="Geometric Data Reader", layout="wide")

def parse_data(file_content):
    """Parse data from file content string"""
    if isinstance(file_content, bytes):
        lines = file_content.decode('utf-8').splitlines()
    else:
        lines = file_content.splitlines()
    
    # Create list to store data
    data = []
    
    for line in lines:
        # Skip empty lines
        if not line.strip():
            continue
            
        # Split the line by semicolons
        parts = line.strip().split(';')
        
        if len(parts) < 2:
            continue
        
        # Extract ID and object type
        obj_id = parts[0]
        obj_type = parts[1]
        
        # Initialize a dictionary for the row
        row = {'ID': obj_id, 'Type': obj_type}
        
        # Extract remaining values (keep empty strings to maintain column mapping)
        remaining_values = [p.strip() for p in parts[2:]]
        
        # Determine column names based on object type
        if obj_type == 'PLANE':
            cols = ['Method', 'X', 'Y', 'Z', 'A', 'B', 'C', '', 'D', 'Dev']
        elif obj_type == 'CIRCLE':
            cols = ['Method', 'X', 'Y', 'Z', 'I', 'J', 'K', '', 'Radius', 'Dev']
        elif obj_type == 'PT-COMP':
            cols = ['Method', 'X', 'Y', 'Z']
        elif obj_type == 'DISTANCE':
            cols = ['', 'X', 'Y', 'Z', '', '', '', '', 'Distance']
        elif obj_type == 'CONE':
            cols = ['Method', 'X', 'Y', 'Z', 'I', 'J', 'K', '', 'Half-Angle', 'Dev']
        elif obj_type == 'INT-CIRCLE':
            cols = ['', 'X', 'Y', 'Z', 'I', 'J', 'K', '', 'Radius']
        elif obj_type == 'SYM-POINT':
            cols = ['', 'X', 'Y', 'Z']
        else:
            # Generic handling for unknown types
            cols = [f'Val{i+1}' for i in range(len(remaining_values))]
        
        # Add values to the row dictionary
        for i, val in enumerate(remaining_values):
            if i < len(cols) and cols[i] and val:
                try:
                    # Try to convert to float if it's a number
                    numeric_pattern = r'^-?\d+\.?\d*$'
                    if re.match(numeric_pattern, val):
                        row[cols[i]] = float(val)
                    else:
                        row[cols[i]] = val
                except:
                    row[cols[i]] = val
        
        data.append(row)
    
    return data

def extract_target_values(data):
    """Extract the specific values we want to copy to Excel"""
    distance_items = []
    int_circle_values = []
    
    # Collect all DISTANCE and INT-CIRCLE items
    for item in data:
        if item['Type'] == 'DISTANCE':
            if 'X' in item and 'ID' in item:
                distance_items.append({
                    'id': item['ID'], 
                    'value': round(abs(item['X']), 2)
                })
        elif item['Type'] == 'INT-CIRCLE':
            # For INT-CIRCLE, we want the Radius column value (not K)
            if 'Radius' in item:
                int_circle_values.append(round(item['Radius'], 2))
    
    # Sort DISTANCE values in the specific order: ID 3, 2, 1, 4
    distance_order = ['3', '2', '1', '4']
    distance_values = []
    
    for target_id in distance_order:
        for item in distance_items:
            if str(item['id']) == target_id:
                distance_values.append(item['value'])
                break
    
    return distance_values, int_circle_values

def extract_lot_prefix(lot_number):
    """Extract LOT prefix from LOT number string
    Example: LOT234(234-245) -> LOT234
             LOT450(450-457) -> LOT450
    """
    import re
    if not lot_number:
        return None
    
    # Match alphabetic letters followed by numbers at the beginning
    match = re.match(r'^([A-Za-z]+\d+)', lot_number)
    if match:
        return match.group(1)
    return None

def validate_date_format(date_string):
    """Validate date format YYYY/MM/DD"""
    import re
    if not date_string:
        return False
    
    # Match YYYY/MM/DD format
    pattern = r'^\d{4}/\d{2}/\d{2}
    """Update Excel file with the extracted values"""
    try:
        # Load the workbook
        wb = load_workbook(excel_file)
        
        # Show available sheets
        st.info(f"📋 Available sheets: {', '.join(wb.sheetnames)}")
        
        # Get the specified sheet or active sheet
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            st.success(f"✓ Using sheet: '{sheet_name}'")
        else:
            ws = wb.active
            st.info(f"✓ Using active sheet: '{ws.title}'")
        
        def write_to_cell(worksheet, cell_ref, value):
            """Write to a cell, handling merged cells"""
            try:
                from openpyxl.utils import get_column_letter
                cell = worksheet[cell_ref]
                
                # Check if cell is a merged cell
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    # Find the merged range that contains this cell
                    for merged_range in worksheet.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            # Get the top-left cell of the merged range
                            min_col, min_row, max_col, max_row = merged_range.bounds
                            top_left_cell = worksheet.cell(row=min_row, column=min_col)
                            top_left_cell.value = value
                            return True
                else:
                    # Normal cell, just write the value
                    cell.value = value
                    return True
            except Exception as e:
                st.error(f"❌ Error updating cell {cell_ref}: {e}")
                return False
        
        # Track successful writes
        successful_writes = 0
        
        # Write LOT information to B column if provided
        if lot_number:
            if write_to_cell(ws, "B1", lot_number):
                successful_writes += 1
                st.success(f"✓ Wrote LOT番号 '{lot_number}' to cell B1")
        
        if inspection_date:
            if write_to_cell(ws, "B2", inspection_date):
                successful_writes += 1
                st.success(f"✓ Wrote 検査日 '{inspection_date}' to cell B2")
        
        if lot_prefix:
            if write_to_cell(ws, "B3", lot_prefix):
                successful_writes += 1
                st.success(f"✓ Wrote LOTプレフィックス '{lot_prefix}' to cell B3")
        
        # Update DISTANCE values
        for i, (value, cell) in enumerate(zip(distance_values, distance_cells)):
            if cell.strip():  # Only update if cell reference is provided
                if write_to_cell(ws, cell.strip(), value):
                    successful_writes += 1
        
        # Update INT-CIRCLE values
        for i, (value, cell) in enumerate(zip(int_circle_values, int_circle_cells)):
            if cell.strip():  # Only update if cell reference is provided
                if write_to_cell(ws, cell.strip(), value):
                    successful_writes += 1
        
        st.info(f"📝 Total cells updated: {successful_writes}")
        
        # Save to BytesIO object
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    except Exception as e:
        st.error(f"Error processing Excel file: {e}")
        return None

def main():
    st.title("全検箇所測定データをExcelに転記するツール")
    
    # File upload with more detailed instructions
    st.subheader("ステップ１：測定データファイルのアップロード")
    st.write("以下の形式のファイルをアップロードしてください。")
    uploaded_file = st.file_uploader(
        "ファイルを選択", 
        type=["txt", "dat", "csv"],
        help="セミコロン区切りのデータファイルを選択してください"
    )
    
    if uploaded_file is not None:
        # Read and parse the file
        file_content = uploaded_file.read()
        data = parse_data(file_content)
        
        if data:
            st.success(f"✅ {len(data)}件のレコードを読み込みました")
            
            # Extract target values
            distance_values, int_circle_values = extract_target_values(data)
            
            st.info(f"📊 抽出されたデータ: DISTANCE値 {len(distance_values)}件, INT-CIRCLE値 {len(int_circle_values)}件")
            
            # Excel export section
            if distance_values or int_circle_values:
                st.subheader("ステップ２：アップロードしたエクセルファイルの指定セルにデータを出力します。")
                
                # Upload Excel file
                excel_file = st.file_uploader(
                    "エクセルファイルを選択",
                    type=["xlsx", "xls"],
                    help="更新したいExcelファイルを選択してください"
                )
                
                if excel_file is not None:
                    st.subheader("ステップ３：セルの指定（オプション）")
                    
                    # Option to choose between automatic A column or custom cells
                    location_option = st.radio(
                        "データを移行するセルの指定:",
                        ["デフォルト設定 (A列に出力)", "カスタム指定"],
                        index=0
                    )
                    
                    if location_option == "デフォルト設定 (A列に出力)":
                        # Automatically set cells to A1-A6
                        distance_cells = [f"A{i+1}" for i in range(len(distance_values))]
                        int_circle_cells = [f"A{i+1+len(distance_values)}" for i in range(len(int_circle_values))]
                        
                        has_distance_cells = True
                        has_int_circle_cells = True
                        
                    else:
                        # Custom cell input (original functionality)
                        st.write("Enter the cell references where you want to paste the values (e.g., A1, B2, C3):")
                        
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            st.write("**DISTANCE Values Cell References:**")
                            distance_cells = []
                            for i in range(len(distance_values)):
                                cell = st.text_input(
                                    f"Cell for DISTANCE value {i+1} ({distance_values[i]})",
                                    key=f"distance_cell_{i}",
                                    placeholder=f"e.g., A{i+1}"
                                )
                                distance_cells.append(cell)
                        
                        with col2:
                            st.write("**INT-CIRCLE Values Cell References:**")
                            int_circle_cells = []
                            for i in range(len(int_circle_values)):
                                cell = st.text_input(
                                    f"Cell for INT-CIRCLE value {i+1} ({int_circle_values[i]})",
                                    key=f"int_circle_cell_{i}",
                                    placeholder=f"e.g., A{i+1+len(distance_values)}"
                                )
                                int_circle_cells.append(cell)
                        
                        # Check if at least one cell reference is provided
                        has_distance_cells = any(cell.strip() for cell in distance_cells)
                        has_int_circle_cells = any(cell.strip() for cell in int_circle_cells)
                    
                    # LOT Information Input Section
                    st.subheader("ステップ４：LOT情報の入力")
                    st.write("エクセルファイルのB列に出力されるLOT情報を入力してください。")
                    
                    with st.form(key="lot_info_form"):
                        lot_number = st.text_input(
                            "LOT番号を入力してください",
                            placeholder="例: LOT234(234-245)",
                            help="LOT番号を入力してください（例: LOT234(234-245)）"
                        )
                        
                        inspection_date = st.text_input(
                            "検査日を入力してください (YYYY/MM/DD)",
                            placeholder="例: 2025/10/07",
                            help="検査日を YYYY/MM/DD 形式で入力してください"
                        )
                        
                        # Show preview of extracted LOT prefix
                        if lot_number:
                            lot_prefix = extract_lot_prefix(lot_number)
                            if lot_prefix:
                                st.info(f"🔍 自動抽出されたLOTプレフィックス: **{lot_prefix}** (B3に出力されます)")
                            else:
                                st.warning("⚠️ LOTプレフィックスを抽出できませんでした")
                        
                        st.write("**出力先:**")
                        st.write("• LOT番号 → **B1**")
                        st.write("• 検査日 → **B2**")
                        st.write("• LOTプレフィックス（自動抽出） → **B3**")
                        
                        submit_button = st.form_submit_button("✓ 確認", type="secondary")
                    
                    # Show confirmation if form is submitted
                    if submit_button:
                        if lot_number and inspection_date:
                            # Validate date format
                            if not validate_date_format(inspection_date):
                                st.error("❌ 検査日は YYYY/MM/DD 形式で入力してください（例: 2025/10/07）")
                            else:
                                lot_prefix = extract_lot_prefix(lot_number)
                                if not lot_prefix:
                                    st.warning("⚠️ LOTプレフィックスを抽出できませんでした。LOT番号を確認してください。")
                                
                                st.success("✅ LOT情報が入力されました！下のボタンをクリックしてエクセルファイルを更新してください。")
                                
                                # Store in session state
                                st.session_state['lot_number'] = lot_number
                                st.session_state['inspection_date'] = inspection_date
                                st.session_state['lot_prefix'] = lot_prefix
                        else:
                            st.error("❌ LOT番号と検査日の両方を入力してください。")
                    
                    # Retrieve from session state if available
                    if 'lot_number' not in st.session_state:
                        st.session_state['lot_number'] = None
                        st.session_state['inspection_date'] = None
                        st.session_state['lot_prefix'] = None
                    
                    # Update Excel button
                    if st.button("📊 エクセルファイルの更新", type="primary"):
                        # Check if LOT information is provided
                        if not st.session_state.get('lot_number') or not st.session_state.get('inspection_date'):
                            st.error("❌ LOT情報を入力してから更新ボタンをクリックしてください。")
                        elif (location_option == "デフォルト設定 (A列に出力)") or (has_distance_cells or has_int_circle_cells):
                            with st.spinner("エクセルファイルを更新中..."):
                                # Always use "sheet" as the sheet name
                                updated_excel = update_excel_file(
                                    excel_file, 
                                    distance_values, 
                                    int_circle_values, 
                                    distance_cells, 
                                    int_circle_cells,
                                    "sheet",  # Fixed sheet name
                                    st.session_state['lot_number'],
                                    st.session_state['inspection_date'],
                                    st.session_state['lot_prefix']
                                )
                            
                            if updated_excel:
                                st.success("✅ エクセルファイルの更新が完了しました!")
                                
                                # Generate dynamic filename
                                # Format: 水平ノズル + {B1} + 全箇所測定 + {B2}
                                lot_num = st.session_state['lot_number']
                                insp_date = st.session_state['inspection_date']
                                dynamic_filename = f"水平ノズル{lot_num}全箇所測定{insp_date}.xlsx"
                                
                                # Download button
                                st.download_button(
                                    label="💾 更新したエクセルファイルのダウンロード",
                                    data=updated_excel.getvalue(),
                                    file_name=dynamic_filename,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                        else:
                            st.warning("⚠️ カスタム指定の場合は、少なくとも1つのセル参照を入力してください。")
        else:
            st.error("No valid data found in the uploaded file")
    else:
        st.info("👆 Browse File ボタンを押して処理するテキストデータをアップロードしてください")

if __name__ == "__main__":
    main()

    return bool(re.match(pattern, date_string))

def update_excel_file(excel_file, distance_values, int_circle_values, distance_cells, int_circle_cells, sheet_name=None, lot_number=None, inspection_date=None, lot_prefix=None):
    """Update Excel file with the extracted values"""
    try:
        # Load the workbook
        wb = load_workbook(excel_file)
        
        # Get the specified sheet or default to "sheet"
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        elif "sheet" in wb.sheetnames:
            ws = wb["sheet"]
            st.info(f"✓ Using sheet: 'sheet'")
        else:
            ws = wb.active
            st.info(f"✓ Using active sheet: '{ws.title}'")
        
        def write_to_cell(worksheet, cell_ref, value):
            """Write to a cell, handling merged cells"""
            try:
                from openpyxl.utils import get_column_letter
                cell = worksheet[cell_ref]
                
                # Check if cell is a merged cell
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    # Find the merged range that contains this cell
                    for merged_range in worksheet.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            # Get the top-left cell of the merged range
                            min_col, min_row, max_col, max_row = merged_range.bounds
                            top_left_cell = worksheet.cell(row=min_row, column=min_col)
                            top_left_cell.value = value
                            st.warning(f"⚠️ Cell {cell_ref} is merged. Wrote {value} to top-left cell {top_left_cell.coordinate}")
                            return True
                else:
                    # Normal cell, just write the value
                    worksheet[cell_ref] = value
                    st.success(f"✓ Wrote {value} to cell {cell_ref}")
                    return True
            except Exception as e:
                st.error(f"❌ Error updating cell {cell_ref}: {e}")
                return False
        
        # Track successful writes
        successful_writes = 0
        
        # Write LOT information to B column if provided
        if lot_number:
            st.write("**Writing LOT information...**")
            if write_to_cell(ws, "B1", lot_number):
                successful_writes += 1
        
        if inspection_date:
            if write_to_cell(ws, "B2", inspection_date):
                successful_writes += 1
        
        if lot_prefix:
            if write_to_cell(ws, "B3", lot_prefix):
                successful_writes += 1
        
        # Update DISTANCE values
        st.write("**Writing DISTANCE values...**")
        for i, (value, cell) in enumerate(zip(distance_values, distance_cells)):
            if cell.strip():  # Only update if cell reference is provided
                if write_to_cell(ws, cell.strip(), value):
                    successful_writes += 1
        
        # Update INT-CIRCLE values
        st.write("**Writing INT-CIRCLE values...**")
        for i, (value, cell) in enumerate(zip(int_circle_values, int_circle_cells)):
            if cell.strip():  # Only update if cell reference is provided
                if write_to_cell(ws, cell.strip(), value):
                    successful_writes += 1
        
        st.info(f"📝 Total successful writes: {successful_writes}")
        
        # Save to BytesIO object
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    except Exception as e:
        st.error(f"Error processing Excel file: {e}")
        return None

def main():
    st.title("全検箇所測定データをExcelに転記するツール")
    
    # File upload with more detailed instructions
    st.subheader("ステップ１：測定データファイルのアップロード")
    st.write("以下の形式のファイルをアップロードしてください。")
    uploaded_file = st.file_uploader(
        "ファイルを選択", 
        type=["txt", "dat", "csv"],
        help="セミコロン区切りのデータファイルを選択してください"
    )
    
    if uploaded_file is not None:
        # Read and parse the file
        file_content = uploaded_file.read()
        data = parse_data(file_content)
        
        if data:
            st.success(f"Successfully loaded {len(data)} records from {uploaded_file.name}")
            
            # Convert to DataFrame
            df = pd.DataFrame(data)
            
            # Show all data
            st.subheader("全ての測定データ")
            st.dataframe(df, use_container_width=True)
            
            # Extract target values
            distance_values, int_circle_values = extract_target_values(data)
            
            # Show extracted values
            col1, col2 = st.columns(2)
            
            with col1:
                st.subheader("📏 距離データ")
                if distance_values:
                    for i, val in enumerate(distance_values, 1):
                        st.write(f"Value {i}: **{val}**")
                else:
                    st.write("No DISTANCE values found")
            
            with col2:
                st.subheader("🔵 円の外周データ")
                if int_circle_values:
                    for i, val in enumerate(int_circle_values, 1):
                        st.write(f"Value {i}: **{val}**")
                else:
                    st.write("No INT-CIRCLE values found")
            
            # Excel export section
            if distance_values or int_circle_values:
                st.subheader("ステップ２：アップロードしたエクセルファイルの指定セルにデータを出力します。")
                
                # Upload Excel file
                excel_file = st.file_uploader(
                    "エクセルファイルを選択",
                    type=["xlsx", "xls"],
                    help="更新したいExcelファイルを選択してください"
                )
                
                if excel_file is not None:
                    st.subheader("ステップ３：セルの指定（オプション）")
                    
                    # Sheet selection
                    try:
                        wb_temp = load_workbook(excel_file)
                        sheet_names = wb_temp.sheetnames
                        wb_temp.close()
                        
                        # Default to "sheet" if it exists, otherwise use first sheet
                        default_index = 0
                        if "sheet" in sheet_names:
                            default_index = sheet_names.index("sheet")
                        
                        selected_sheet = st.selectbox(
                            "シートを選択:",
                            options=sheet_names,
                            index=default_index,
                            help="データを書き込むシートを選択してください"
                        )
                        
                        if selected_sheet == "sheet":
                            st.success("✓ 'sheet' シートが選択されています")
                    except Exception as e:
                        st.error(f"Error reading Excel sheets: {e}")
                        selected_sheet = None
                    
                    # Option to choose between automatic A column or custom cells
                    location_option = st.radio(
                        "データを移行するセルの指定:",
                        ["デフォルト設定", "指定する場合"],
                        index=0
                    )
                    
                    if location_option == "デフォルト設定":
                        # Automatically set cells to A1-A6
                        distance_cells = [f"A{i+1}" for i in range(len(distance_values))]
                        int_circle_cells = [f"A{i+1+len(distance_values)}" for i in range(len(int_circle_values))]
                        
                        st.write("**Values will be placed in:**")
                        for i, val in enumerate(distance_values):
                            st.write(f"• DISTANCE value {val} → **A{i+1}**")
                        for i, val in enumerate(int_circle_values):
                            st.write(f"• INT-CIRCLE value {val} → **A{i+1+len(distance_values)}**")
                        
                        has_distance_cells = True
                        has_int_circle_cells = True
                        
                    else:
                        # Custom cell input (original functionality)
                        st.write("Enter the cell references where you want to paste the values (e.g., A1, B2, C3):")
                        
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            st.write("**DISTANCE Values Cell References:**")
                            distance_cells = []
                            for i in range(len(distance_values)):
                                cell = st.text_input(
                                    f"Cell for DISTANCE value {i+1} ({distance_values[i]})",
                                    key=f"distance_cell_{i}",
                                    placeholder=f"e.g., A{i+1}"
                                )
                                distance_cells.append(cell)
                        
                        with col2:
                            st.write("**INT-CIRCLE Values Cell References:**")
                            int_circle_cells = []
                            for i in range(len(int_circle_values)):
                                cell = st.text_input(
                                    f"Cell for INT-CIRCLE value {i+1} ({int_circle_values[i]})",
                                    key=f"int_circle_cell_{i}",
                                    placeholder=f"e.g., A{i+1+len(distance_values)}"
                                )
                                int_circle_cells.append(cell)
                        
                        # Check if at least one cell reference is provided
                        has_distance_cells = any(cell.strip() for cell in distance_cells)
                        has_int_circle_cells = any(cell.strip() for cell in int_circle_cells)
                    
                    # LOT Information Input Section
                    st.subheader("ステップ４：LOT情報の入力")
                    st.write("エクセルファイルのB列に出力されるLOT情報を入力してください。")
                    
                    with st.form(key="lot_info_form"):
                        lot_number = st.text_input(
                            "LOT番号を入力してください",
                            placeholder="例: LOT234(234-245)",
                            help="LOT番号を入力してください（例: LOT234(234-245)）"
                        )
                        
                        inspection_date = st.text_input(
                            "検査日を入力してください (YYYY/MM/DD)",
                            placeholder="例: 2025/10/07",
                            help="検査日を YYYY/MM/DD 形式で入力してください"
                        )
                        
                        # Show preview of extracted LOT prefix
                        if lot_number:
                            lot_prefix = extract_lot_prefix(lot_number)
                            if lot_prefix:
                                st.info(f"🔍 自動抽出されたLOTプレフィックス: **{lot_prefix}** (B3に出力されます)")
                            else:
                                st.warning("⚠️ LOTプレフィックスを抽出できませんでした")
                        
                        st.write("**出力先:**")
                        st.write("• LOT番号 → **B1**")
                        st.write("• 検査日 → **B2**")
                        st.write("• LOTプレフィックス（自動抽出） → **B3**")
                        
                        submit_button = st.form_submit_button("✓ 確認", type="secondary")
                    
                    # Show confirmation if form is submitted
                    if submit_button:
                        if lot_number and inspection_date:
                            lot_prefix = extract_lot_prefix(lot_number)
                            st.success("✅ LOT情報が入力されました！下のボタンをクリックしてエクセルファイルを更新してください。")
                            
                            # Store in session state
                            st.session_state['lot_number'] = lot_number
                            st.session_state['inspection_date'] = inspection_date
                            st.session_state['lot_prefix'] = lot_prefix
                        else:
                            st.error("❌ LOT番号と検査日の両方を入力してください。")
                    
                    # Retrieve from session state if available
                    if 'lot_number' not in st.session_state:
                        st.session_state['lot_number'] = None
                        st.session_state['inspection_date'] = None
                        st.session_state['lot_prefix'] = None
                    
                    # Update Excel button
                    if st.button("📊 エクセルファイルの更新", type="primary"):
                        # Check if LOT information is provided
                        if not st.session_state.get('lot_number') or not st.session_state.get('inspection_date'):
                            st.error("❌ LOT情報を入力してから更新ボタンをクリックしてください。")
                        elif (location_option == "デフォルト設定") or (has_distance_cells or has_int_circle_cells):
                            with st.spinner("Updating Excel file..."):
                                updated_excel = update_excel_file(
                                    excel_file, 
                                    distance_values, 
                                    int_circle_values, 
                                    distance_cells, 
                                    int_circle_cells,
                                    selected_sheet,
                                    st.session_state['lot_number'],
                                    st.session_state['inspection_date'],
                                    st.session_state['lot_prefix']
                                )
                            
                            if updated_excel:
                                st.success("✅ エクセルファイルの更新が完了しました!")
                                
                                # Download button
                                st.download_button(
                                    label="💾 更新したエクセルファイルのダウンロード",
                                    data=updated_excel.getvalue(),
                                    file_name=f"updated_{excel_file.name}",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                                
                                # Show summary
                                st.subheader("更新内容の概要")
                                
                                # LOT information summary
                                if st.session_state.get('lot_number'):
                                    st.write("**LOT情報:**")
                                    st.write(f"✓ LOT番号 {st.session_state['lot_number']} → Cell B1")
                                if st.session_state.get('inspection_date'):
                                    st.write(f"✓ 検査日 {st.session_state['inspection_date']} → Cell B2")
                                if st.session_state.get('lot_prefix'):
                                    st.write(f"✓ LOTプレフィックス {st.session_state['lot_prefix']} → Cell B3")
                                
                                st.write("**測定データ:**")
                                for i, (val, cell) in enumerate(zip(distance_values, distance_cells)):
                                    if cell.strip():
                                        st.write(f"✓ DISTANCE value {val} → Cell {cell}")
                                
                                for i, (val, cell) in enumerate(zip(int_circle_values, int_circle_cells)):
                                    if cell.strip():
                                        st.write(f"✓ INT-CIRCLE value {val} → Cell {cell}")
                        else:
                            st.warning("⚠️ Please provide at least one cell reference in custom mode.")
        else:
            st.error("No valid data found in the uploaded file")
    else:
        st.info("👆 Browse File ボタンを押して処理するテキストデータをアップロードしてください")

if __name__ == "__main__":
    main()